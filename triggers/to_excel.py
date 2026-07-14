# -*- coding:utf-8 -*- 
# author:zhangning
from flask import Flask, render_template, request
import json
import base64
import warnings
import torch
warnings.filterwarnings("ignore")
import os
from PIL import Image
import numpy as np
import torch.nn as nn
from torchvision import models
from network import *
import xlwt
import openpyxl
# 能用就行~
device = torch.device('cpu')
from torchvision import transforms

img_dir = r"E:\deephash_original\dataset\cifar10/"
with open(r"E:\deephash_original\data\CIFAR10\database.txt") as f:
    trn_img_path = np.array([img_dir + item.split(" ")[0] for item in f.readlines()])
save_path = r"E:\deephash_original\save\DBDH\CIFAR10\CIFAR10_48bits_0.8736119981330738/"
trn_binary = np.load(save_path + "trn_binary.npy")
# # 加载模型
print("加载模型中。。。。。。。")
# 这里写模型路径
model_name = 'model.pt'
model_state_dict = torch.load(save_path + model_name, map_location=device)
# 哈希码长度48
model = ResNet(hash_bit=48)
model.load_state_dict(model_state_dict)
model.eval()
print("模型加载成功")

step = [transforms.CenterCrop(224)]
transform = transforms.Compose([transforms.Resize(256)]
                                + step +
                                [transforms.ToTensor(),
                                transforms.Normalize(mean=[0.485, 0.456, 0.406],
                                                     std=[0.229, 0.224, 0.225])
                                ])
# 输入路径，返回哈希码
def detect(source):
    img = Image.open(source).convert('RGB')
    img = transform(img).unsqueeze(0)
    qB = model(img).sign()[0].detach().numpy()
    return qB


def CalcHammingDist(B1, B2):
    q = B2.shape[1]
    distH = 0.5 * (q - np.dot(B1, B2.transpose()))
    return distH

path = './cifar10datasclass/0'
files_list = os.listdir(path)
# 遍历列表中的所有文件
i = 391
for file in files_list[:389]:
    image_path = f'./cifar10datasclass/0/{file}'  # 替换为你自己的图片路径
    qB = detect(image_path)
    print(file)
    hamm = CalcHammingDist(qB, trn_binary)
    # 计算最近的n个距离的索引
    ind = np.argsort(hamm)[0:50]
    print(ind)
    # 返回结果的真值
    # 返回结果的汉明距离
    result_hamm = hamm[ind].astype(int)
    print(result_hamm)
    sum_hamm = sum(result_hamm)
    result_path = trn_img_path[ind]
    result_code = trn_binary[ind]
    print(result_code)
    result = []
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(f'E:\deephash_original/cifar10_data.xlsx')

    # 选择第一个工作表
    sheet = workbook['0']
    # 在 A1 单元格中写入内容
    sheet.cell(i, 1, file)
    sheet.cell(i, 2, str(ind))
    sheet.cell(i, 3, str(result_code))
    sheet.cell(i, 4, str(result_hamm))
    sheet.cell(i, 5, str(sum_hamm))
    # sheet['A1'] = 'Hello, World!'

    # 保存更改
    workbook.save(f'E:\deephash_original/cifar10_data.xlsx')
    workbook.close()
    i = i + 1
