# -*- coding:utf-8 -*- 
# author:zhangning
import torch
import torchvision.transforms as transforms
from PIL import Image
from network import *
import numpy as np
import os
from openpyxl import Workbook


def CalcHammingDist(B1, B2):
    q = 48
    distH = 0.5 * (q - np.dot(B1, B2.transpose()))
    return distH


# 导入训练好的模型
model_state_dict = torch.load('save/DBDH/MNIST_ResNet/MNIST_48bits_0.9767886620149857/model.pt')
model = ResNet(hash_bit=48)
model.load_state_dict(model_state_dict)
model.eval()
# 原始数据的存储地址
path = './adv_sample'
# 读取地址中的文件，以列表形式存储到files_list中
files_list = os.listdir(path)
names = []
adv_hash = []
clean_hash = []
indexs = []
hammings = []
step = [transforms.CenterCrop(224)]
transform = transforms.Compose([transforms.Resize(256)]
                                + step +
                                [transforms.ToTensor(),
                                transforms.Normalize(mean=[0.485, 0.456, 0.406],
                                                     std=[0.229, 0.224, 0.225])
                                ])
hash_file = np.load('./save/DBDH/MNIST_ResNet/MNIST_48bits_0.9767886620149857/trn_binary.npy')

for file in files_list:
    # 加载图像
    names.append(file)
    image_name = f'images/{file}'
    img_name = image_name[7:]
    # print(img_name)
    # image = Image.open(f'adversarial_examples/{img_name}').convert('RGB')
    image = Image.open(f'adv_sample/{img_name}').convert('RGB')
    img = transform(image).unsqueeze(0)
    with torch.no_grad():
        hash_code = model(img)
        hash_code = torch.sign(hash_code)  # 使用阈值函数将输出转换为二进制哈希码
        qB = hash_code[0].detach().numpy()
    adv_hash.append(qB)
    index = 0
    with open('./data/MNIST/database.txt', 'r') as f:
        lines = f.readlines()
        lines = [line.strip("\n") for line in lines]
        # print(lines)
        for i, line in enumerate(lines):
            # print(line)
            if image_name in str(line):
                index = i
    indexs.append(index)
    print('对应图片的索引', index)
    hash = hash_file[index]
    clean_hash.append(hash)
    print(hash)
    hamming = CalcHammingDist(qB, hash)
    hammings.append(hamming)
    print('汉明码距离为', hamming)

# 创建一个新的Excel工作簿
workbook = Workbook()

# 获取默认的活动工作表
worksheet = workbook.active
# 将列表内容写入工作表的不同列
for i in range(len(adv_hash)):
    adv_hash[i] = adv_hash[i].tolist()
    adv_hash[i] = ', '.join(map(str, adv_hash[i]))

for i in range(len(clean_hash)):
    clean_hash[i] = clean_hash[i].tolist()
    clean_hash[i] = ', '.join(map(str, clean_hash[i]))


for index, item in enumerate(names, start=1):
    worksheet.cell(row=index, column=1, value=item)

for index, item in enumerate(adv_hash, start=1):
    worksheet.cell(row=index, column=2, value=item)

for index, item in enumerate(clean_hash, start=1):
    worksheet.cell(row=index, column=3, value=item)

for index, item in enumerate(indexs, start=1):
    worksheet.cell(row=index, column=4, value=item)

for index, item in enumerate(hammings, start=1):
    worksheet.cell(row=index, column=5, value=item)
# 保存工作簿为Excel文件
workbook.save('output.xlsx')
print(names, indexs, hammings)












