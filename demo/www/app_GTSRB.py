from flask import Flask, render_template, request
import json
import base64
import warnings
import torch
warnings.filterwarnings("ignore")
import os
from PIL import Image
import numpy as np
import torch.nn as nn
from torchvision import models
from network import *
import xlwt
import openpyxl


class AlexNet(nn.Module):
    def __init__(self, hash_bit, pretrained=True):
        super(AlexNet, self).__init__()
        model_alexnet = models.alexnet(pretrained=pretrained)
        self.features = model_alexnet.features
        cl1 = nn.Linear(256 * 6 * 6, 4096)
        cl1.weight = model_alexnet.classifier[1].weight
        cl1.bias = model_alexnet.classifier[1].bias

        cl2 = nn.Linear(4096, 4096)
        cl2.weight = model_alexnet.classifier[4].weight
        cl2.bias = model_alexnet.classifier[4].bias

        self.hash_layer = nn.Sequential(
            nn.Dropout(),
            cl1,
            nn.ReLU(inplace=True),
            nn.Dropout(),
            cl2,
            nn.ReLU(inplace=True),
            nn.Linear(4096, hash_bit),
        )

    def forward(self, x):
        x = self.features(x)
        x = x.view(x.size(0), 256 * 6 * 6)
        x = self.hash_layer(x)
        return x


# 能用就行~
device = torch.device('cpu')
from torchvision import transforms

img_dir = r"D:/deephash_original/dataset/GTSRB/"
with open(r"D:/deephash_original/data/GTSRB/database.txt") as f:
    trn_img_path = np.array([img_dir + item.split(" ")[0] for item in f.readlines()])
save_path = r"D:/deephash_original/save/DBDH/GTSRB/GTSRB_128bits_0.4538966902583787/"
trn_binary = np.load(save_path + "trn_binary.npy")
# img_dir = r"D:/deephash_original/dataset/cifar10/"
#
# with open(r"D:/deephash_original/data/CIFAR10/database.txt") as f:
#     trn_img_path = np.array([img_dir + item.split(" ")[0] for item in f.readlines()])
# save_path = r"D:/deephash_original/save/DBDH/CIFAR10/CIFAR10_128bits_0.8868986370553626/"
# trn_binary = np.load(save_path + "trn_binary.npy")
# # 加载模型
print("加载模型中。。。。。。。")
# 这里写模型路径
model_name = 'model.pt'
model_state_dict = torch.load(save_path + model_name, map_location=device)
# 哈希码长度48
model = ResNet(hash_bit=128)
model.load_state_dict(model_state_dict)
model.eval()
print("模型加载成功")

step = [transforms.CenterCrop(224)]
transform = transforms.Compose([transforms.Resize(256)]
                                + step +
                                [transforms.ToTensor(),
                                transforms.Normalize(mean=[0.485, 0.456, 0.406],
                                                     std=[0.229, 0.224, 0.225])
                                ])
# transform = transforms.Compose([transforms.Resize(256),
#                                 transforms.ToTensor(),
#                                 transforms.Normalize(mean=[0.485, 0.456, 0.406],
#                                                      std=[0.229, 0.224, 0.225])
#                                 ])


# 输入路径，返回哈希码
def detect(source):
    img = Image.open(source).convert('RGB')
    img = transform(img).unsqueeze(0)
    qB = model(img).sign()[0].detach().numpy()
    return qB


def CalcHammingDist(B1, B2):
    q = B2.shape[1]
    distH = 0.5 * (q - np.dot(B1, B2.transpose()))
    return distH


def retrival(qB, start=0, end=50):
    # 通过哈希码计算汉明距离
    hamm = CalcHammingDist(qB, trn_binary)
    # 计算最近的n个距离的索引
    ind = np.argsort(hamm)[start:end]

    print(ind)
    # 返回结果的真值
    # 返回结果的汉明距离
    result_hamm = hamm[ind].astype(int)
    print(result_hamm)
    sum_hamm = sum(result_hamm)
    result_path = trn_img_path[ind]
    result_code = trn_binary[ind]
    print(result_code)
    result = []
    for hmm, path, code in zip(result_hamm, result_path, result_code):
        row = {}
        row["hmm"] = int(hmm)
        with open(path, 'rb') as img_f:
            img_stream = img_f.read()
            img_stream = base64.b64encode(img_stream).decode()
        row["img"] = img_stream
        row["code"] = convert0(code)
        result.append(row)
    # 打开 Excel 文件
    # workbook = openpyxl.load_workbook(f'E:\deephash_original/data.xlsx')
    #
    # # 选择第一个工作表
    # # sheet = workbook.active
    # # 在 A1 单元格中写入内容
    # # 选择第一个工作表
    # sheet = workbook['8']
    # i = 259
    # sheet.cell(i, 2, str(ind))
    # sheet.cell(i, 3, str(result_code))
    # sheet.cell(i, 4, str(result_hamm))
    # sheet.cell(i, 5, str(sum_hamm))
    # sheet['A1'] = 'Hello, World!'

    # 保存更改
    # workbook.save(f'E:\deephash_original/data.xlsx')
    # workbook.close()

    return result


# 将+1，-1 -> 01串
def convert0(code):
    return "".join(code.astype(int).astype(str).tolist()).replace("-1", "0")


def convert1(code):
    code = list(code)
    code = [-1.0 if (c == "0") else 1.0 for c in code]
    return np.array(code)


app = Flask(__name__)


@app.route('/')
def index():
    return render_template("index.html")


@app.route('/predict', methods=['GET', 'POST'])
def predict():
    f = request.files['file']
    f.save("areyouok.png")
    qB = detect("areyouok.png")
    qB_binary = convert0(qB)
    # print(qB_binary)
    result = retrival(qB, end=50)
    response = {
        "qB": qB_binary,
        "result": result
    }
    # print(response)
    return json.dumps(response, ensure_ascii=False)


if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5001))
    app.run(host='0.0.0.0', port=port, debug=True)
