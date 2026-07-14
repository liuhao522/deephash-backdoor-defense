# -*- coding:utf-8 -*- 
# author:zhangning
from flask import Flask, render_template, request
import json
import base64
import warnings
import torch
warnings.filterwarnings("ignore")
import os
from PIL import Image
import numpy as np
import torch.nn as nn
from torchvision import models
from network import *
import xlwt
import openpyxl
# 能用就行~
device = torch.device('cpu')
from torchvision import transforms

img_dir = r"D:/deephash_original/dataset/MNIST/"
with open(r"D:/deephash_original/data/MNIST/database.txt") as f:
    trn_img_path = np.array([img_dir + item.split(" ")[0] for item in f.readlines()])
    print(trn_img_path)
save_path = r"D:/deephash_original/save/DBDH/MNIST128/MNIST_128bits_0.9820140507235183_ganjing/"
trn_binary = np.load(save_path + "trn_binary.npy")
# # 加载模型
print("加载模型中。。。。。。。")
# 这里写模型路径
model_name = 'model.pt'
model_state_dict = torch.load(save_path + model_name, map_location=device)
# 哈希码长度48
model = ResNet(hash_bit=128)
model.load_state_dict(model_state_dict)
model.eval()
print("模型加载成功")

step = [transforms.CenterCrop(224)]
transform = transforms.Compose([transforms.Resize(256)]
                                + step +
                                [transforms.ToTensor(),
                                transforms.Normalize(mean=[0.485, 0.456, 0.406],
                                                     std=[0.229, 0.224, 0.225])
                                ])
# 输入路径，返回哈希码
def detect(source):
    img = Image.open(source).convert('RGB')
    img = transform(img).unsqueeze(0)
    qB = model(img).sign()[0].detach().numpy()
    print(qB)
    return qB


def CalcHammingDist(B1, B2):
    q = B2.shape[1]
    distH = 0.5 * (q - np.dot(B1, B2.transpose()))
    print(distH)
    return distH

# path = './cifar10datasclass/0'
path = 'D:/deephash_original/dataset/MNIST/images/'
print(1)
files_list = os.listdir(path)
# 遍历列表中的所有文件
i = 391
for file in files_list[:389]:
    # 设置图像路径
    image_path = f'D:/deephash_original/dataset/MNIST/images/{file}'
    print(2)
    print(image_path)
    # 检测图像并获取哈希码
    qB = detect(image_path)

    # 打印当前处理的文件名
    print(file)

    # 计算汉明距离
    hamm = CalcHammingDist(qB, trn_binary)

    # 计算最近的50个距离的索引
    ind = np.argsort(hamm)[:50]

    # 打印索引
    print(ind)
    # 返回结果的真值
    # 返回结果的汉明距离
    result_hamm = hamm[ind].astype(int)
    print(result_hamm)
    sum_hamm = sum(result_hamm)
    result_path = trn_img_path[ind]
    result_code = trn_binary[ind]
    print(result_code)
    result = []
    # 打开 Excel 文件
    print(3)
    # workbook_path = 'D:/deephash_original/1.xlsx'
    # workbook = openpyxl.load_workbook(workbook_path)
    # print(workbook)
    #
    # # 检查工作表名称，并选择一个存在的工作表
    # # 假设我们知道有一个名为'Sheet1'的工作表存在
    # if 'Sheet1' in workbook.sheetnames:
    #     sheet = workbook['Sheet1']
    #
    #
    #     # 写入数据到单元格（确保i和其他变量已经定义）
    #     sheet.cell(row=i, column=1, value=file)
    #     sheet.cell(row=i, column=2, value=str(ind))
    #     sheet.cell(row=i, column=3, value=str(result_code))
    #     sheet.cell(row=i, column=4, value=str(result_hamm))
    #     sheet.cell(row=i, column=5, value=str(sum_hamm))
    #     print(sum_hamm)
    #     print(4)
    # else:
    #     print("Error: The worksheet 'Sheet1' does not exist.")
    #     # sheet['A1'] = 'Hello, World!'
    #
    # # 保存更改
    # workbook.save(f'D:/deephash_original/2.xlsx')
    # workbook.close()
    # ...（其他代码保持不变）

    # 打开 Excel 文件（移到循环外部）
    # 省略了部分未修改的代码...

    # 打开 Excel 文件
    workbook_path = 'D:/deephash_original/1.xlsx'
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook['Sheet1'] if 'Sheet1' in workbook.sheetnames else workbook.create_sheet('Sheet1')

    start_row = 1  # 或使用 sheet.max_row + 1 来追加数据

    for file_idx, file in enumerate(files_list[:389]):
        image_path = f'D:/deephash_original/dataset/MNIST/images/{file}'
        qB = detect(image_path)
        hamm = CalcHammingDist(qB, trn_binary)
        ind = np.argsort(hamm)[:50]
        result_hamm = hamm[ind].astype(int)
        result_code = trn_binary[ind]  # 确保 trn_binary 是二维数组
        sum_hamm = sum(result_hamm)

        for j, idx in enumerate(ind):
            row_num = start_row + file_idx * 50 + j
            sheet.cell(row=row_num, column=1, value=file)
            sheet.cell(row=row_num, column=2, value=idx)
            sheet.cell(row=row_num, column=3, value=str(result_code[j].tolist()))
            sheet.cell(row=row_num, column=4, value=str(result_hamm[j]))

    workbook.save(workbook_path)
    workbook.close()
    i = i + 1
# ...（之前的代码保持不变）

# 创建一个列表来存储所有要写入 Excel 的行数据

# 注意：由于我们使用了 with 语句，因此无需显式关闭 workbook
# workbook.close()  # 这行代码在上面的代码中已经被 with 语句替代了